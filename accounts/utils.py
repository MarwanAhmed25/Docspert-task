import csv
import io
import json
from .models import Account
from openpyxl import load_workbook

def get_file_data(filename):
    extension = filename.name.split('.')[1]
    if extension == 'csv':
        file = filename.read().decode('utf-8')
        wb = csv.DictReader(io.StringIO(file))
        for obj in wb:
            Account.objects.create(name=obj['Name'], id=obj['ID'], balance=obj['Balance'])

    elif extension == 'xlsx':
        wb = load_workbook(filename=filename).worksheets
        for sheet in wb:
            for row in sheet.iter_rows(min_row=2):
                pk, name, balance = row[0].value, row[1].value, row[2].value
                Account.objects.create(name=name, id=pk, balance=balance)

    elif extension == 'json':
        wb = json.load(filename)
        for obj in wb:
            Account.objects.create(name=obj['Name'], id=obj['ID'], balance=obj['Balance'])
